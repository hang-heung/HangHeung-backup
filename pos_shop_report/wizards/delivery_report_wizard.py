from datetime import datetime, timedelta
from odoo import models, fields, api
from io import BytesIO
import base64
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

class DeliveryReportWizard(models.TransientModel):
    _name = 'delivery.report.wizard'
    _description = 'Delivery XLS Report Wizard'

    date_from = fields.Date(required=True)
    date_to = fields.Date(required=True)
    file = fields.Binary('File', readonly=True)
    file_name = fields.Char('Filename', readonly=True)
    # HH-CUSTOM: filter the export by isolation flag. 'all' keeps every
    # delivery; 'wedding_only' / 'b2b_only' / 'either' restrict to the
    # corresponding 嫁囍單 / B2B單 records carried from the source SO.
    flag_filter = fields.Selection(
        [
            ('all', 'All Orders'),
            ('wedding_only', '嫁囍單 only'),
            ('b2b_only', 'B2B單 only'),
            ('either', '嫁囍單 + B2B單'),
        ],
        string='Order Type Filter',
        default='all',
        required=True,
    )

    def action_generate_report(self):
        # date_to is a Date but scheduled_date is a Datetime; use exclusive
        # next-day upper bound so deliveries scheduled at any time on date_to
        # are still included.
        date_to_excl = self.date_to + timedelta(days=1)
        pickings = self.env['stock.picking'].search([
            ('picking_type_code', '=', 'outgoing'),
            ('scheduled_date', '>=', self.date_from),
            ('scheduled_date', '<', date_to_excl),
            ('state', 'in', ['assigned', 'confirmed', 'waiting', 'done']),
        ])

        # HH-CUSTOM: apply the wedding/B2B flag filter selected on the wizard.
        if self.flag_filter == 'wedding_only':
            pickings = pickings.filtered(lambda p: p.is_wedding_order)
        elif self.flag_filter == 'b2b_only':
            pickings = pickings.filtered(lambda p: p.is_b2b_order)
        elif self.flag_filter == 'either':
            pickings = pickings.filtered(lambda p: p.is_wedding_order or p.is_b2b_order)

        products = self.env['product.product'].browse(
            pickings.move_ids_without_package.product_id.ids
        ).sorted(key=lambda p: p.name)

        wb = Workbook()
        ws = wb.active
        ws.title = "Delivery Report"

        base_headers = ["Dn Date", "Dn No", "Client", "Shop Code", "來源單據", "Origin Chain", "Remark", "嫁囍單", "B2B單"]
        for col_idx, header in enumerate(base_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
        start_col = len(base_headers) + 1
        for idx, product in enumerate(products, start=start_col):
            cell = ws.cell(row=1, column=idx, value=f"{product.default_code or ''}\n{product.name}")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)
            ws.cell(row=2, column=idx, value="Item Qty").alignment = Alignment(horizontal="center")

        ws.column_dimensions[get_column_letter(1)].width = 20   # Dn Date
        ws.column_dimensions[get_column_letter(2)].width = 15   # Dn No
        ws.column_dimensions[get_column_letter(3)].width = 30   # Client
        ws.column_dimensions[get_column_letter(4)].width = 12   # Shop Code
        ws.column_dimensions[get_column_letter(5)].width = 22   # 來源單據
        ws.column_dimensions[get_column_letter(6)].width = 40   # Origin Chain
        ws.column_dimensions[get_column_letter(7)].width = 25   # Remark
        ws.column_dimensions[get_column_letter(8)].width = 10   # 嫁囍單
        ws.column_dimensions[get_column_letter(9)].width = 10   # B2B單
        for col_idx in range(start_col, start_col + len(products)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        row_idx = 3
        for picking in pickings:
            note_text = ""
            if picking.note:
                soup = BeautifulSoup(picking.note, "html.parser")
                note_text = soup.get_text(separator=" ", strip=True)

            shop_code = ""
            if picking.origin and picking.partner_id and picking.partner_id.name:
                shop_code = picking.partner_id.name.split("-")[0].strip()

            row = [
                picking.scheduled_date.strftime('%d/%m/%Y') if picking.scheduled_date else "",
                picking.name or "",
                picking.partner_id.name or "",
                shop_code,
                picking.origin or "",
                picking.full_origin_chain or "",
                note_text,
                "Y" if picking.is_wedding_order else "",
                "Y" if picking.is_b2b_order else "",
            ]

            for product in products:
                # Show picked qty once the move is done; otherwise fall
                # back to demanded qty so non-validated pickings still
                # appear with the "supposed to ship" quantity.
                qty = sum(
                    (line.quantity if line.state == 'done' else line.product_uom_qty)
                    for line in picking.move_ids_without_package
                    if line.product_id == product
                )
                row.append(qty if qty != 0 else "")

            ws.append(row)
            row_idx += 1

        total_row = ["Grand Total", "", "", "", "", "", "", "", ""]
        for idx, product in enumerate(products, start=start_col):
            col_letter = ws.cell(row=2, column=idx).column_letter
            total_row.append(f"=SUM({col_letter}3:{col_letter}{row_idx-1})")
        ws.append(total_row)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        file_data = base64.b64encode(fp.read())
        fp.close()

        date_from_str = self.date_from.strftime('%d-%m-%Y')
        date_to_str = self.date_to.strftime('%d-%m-%Y')
        flag_label = dict(self._fields['flag_filter'].selection).get(
            self.flag_filter, self.flag_filter,
        )
        attachment = self.env['ir.attachment'].create({
            'name': f"DN List Report_{date_from_str}-{date_to_str}-{flag_label}.xlsx",
            'type': 'binary',
            'datas': file_data,
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
